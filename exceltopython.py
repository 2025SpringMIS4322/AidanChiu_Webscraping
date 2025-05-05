import openpyxl as xl

# load in an excel workbook
wb = xl.load_workbook('example.xlsx')

# Create a list of all the sheets in an excel file
sn = wb.sheetnames
print(sn)

# Create sheet object
sheet1 = wb['Sheet1']

# Create cell object
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1)

# Cell object functions
print(cellA1.value) # Return value of a given cell
print(cellA1.row) # Return row # 
print(cellA1.column) # Return column #
print(cellA1.coordinate) # Return cell coordinate i.e. A1

# Print out 'Apples' in cellB1. NOT A ZERO BASED INDEX SYSTEM
print(sheet1.cell(1,2).value)
print(sheet1.max_row) # Return max # of rows with data
print(sheet1.max_column) # Return max # of columns with data

# Print out names of all fruits
for i in range(1, sheet1.max_row+1): # Range() function doesn't include upper limit. Exclusive. Must add +1.
    print(sheet1.cell(i,2).value) # i iterates thru every row

print(xl.utils.get_column_letter(900)) # what is column letter for 900th column in Excel?
print(xl.utils.column_index_from_string('AHF')) # what is column # for a corresponding letter designation?


# 
for currentrow in sheet1['A1':'C3']:
    #print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)



for currentrow in sheet1.iter_rows(min_row = 2): # 'iter_rows() --> Iterates thru rows in excel sheet
    print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)