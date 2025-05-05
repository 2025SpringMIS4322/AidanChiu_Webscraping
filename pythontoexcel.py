import openpyxl as xl
from openpyxl.styles import Font

# Create workbook object
wb = xl.Workbook()

# Create worksheet object
ws = wb.active

# Name your new worksheet
ws.title = 'First Sheet'

# Create new worksheet to workbook
wb.create_sheet(index = 1, title = 'Second Sheet') # index [1] creates sheet after the initial sheet that's autocreated in a workbook

# Write content to First Sheet
ws['A1'] = 'Invoice'

# Edit font of words in cell A1
ws['A1'].font = Font(name = 'Times New Roman', size = 24, bold = True)

# set a font object for easier formatting
font_object = Font(name = 'Times New Roman', size = 24, bold = True)

# write some more stuff to cells in your new worksheet
ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

# Merge cells
ws.merge_cells('A1:B1')

# Unmerge cells
# ws.unmerge_cells('A1:B1')

# MORE TEXT
ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150


# Create 'totals' cell
ws['A8'] = 'Total'
ws['A8'].font = font_object

# Create an equation for cell B8
ws['B8'] = '=SUM(B2:B7)'

# Change width dimensions of column A
ws.column_dimensions['A'].width = 15



# Produce Report Example (Read from another Excel sheet, and write to this one)
write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost per Pound'
write_sheet['C1'] = 'Amt sold'
write_sheet['D1'] = 'Total'

row_counter = 2 # last row part of produce

for row in read_ws.iter_rows(min_row = 2):
    name = row[0].value
    cost = row[1].value
    amt_sold = row[2].value
    total = row[3].value

    #                Row      column
    write_sheet.cell(row_counter,1).value = name
    write_sheet.cell(row_counter,2).value = cost
    write_sheet.cell(row_counter,3).value = amt_sold
    write_sheet.cell(row_counter,4).value = total

    row_counter += 1

# Write in summary rows
summary_row = row_counter + 1 

write_sheet['B' + str(summary_row)] = 'Total'
write_sheet['B' + str(summary_row)].font = Font(size=16, bold=True)

write_sheet['C' + str(summary_row)] = '=SUM(C2:C' + str(row_counter) + ')'
write_sheet['D' + str(summary_row)] = '=SUM(D2:D' + str(row_counter) +')'


summary_row2 = row_counter + 2

write_sheet['B' + str(summary_row2)] = 'Average'
write_sheet['B' + str(summary_row2)].font = Font(size=16, bold=True)

write_sheet['C' + str(summary_row2)] = '=AVERAGE(C2:C' + str(row_counter) + ')'
write_sheet['D' + str(summary_row2)] = '=AVERAGE(D2:D' + str(row_counter) +')'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 16
write_sheet.column_dimensions['C'].width = 16
write_sheet.column_dimensions['D'].width = 16

# Cell number formatting
for cell in write_sheet["C:C"]:
    cell.number_format = "#,##0"

for cell in write_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'


# Save workbook
wb.save('PythonToExcel.xlsx')
