
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2025/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

movie_rows = soup.findAll('tr')
#print(movie_rows[1])

# Create an excel workbook
wb = xl.Workbook()

# Create worksheet in created workbook
ws = wb.active
ws.title = 'Box Office Report'

ws['A1'] = 'Rank'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Theaters'
ws['F1'] = 'Avg / Theater'


# find all 
for x in range(1, 6):
    td = movie_rows[x].findAll('td')
    rank = td[0].text
    title = td[1].text
    gross = int(td[5].text.strip('$').replace(',', '')) # might need to account for hidden rows in 'inspection mode' sometimes when scraping
    theaters = int(td[6].text.replace(',', ''))
    release_date = td[8].text

    avg_per_theater = round(gross / theaters, 2)
    #print(avg_per_theater)

    # Write results to Excel workbook/worksheet

    ws['A' + str(x+1)] = rank # 'x' acts as counter
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = release_date
    ws['D' + str(x+1)] = gross
    ws['E' + str(x+1)] = theaters
    ws['F' + str(x+1)] = avg_per_theater

ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 26

header_font = Font(size=16, bold=True)

for cell in ws[1:1]: # iterate thru first row
    cell.font = header_font

# Formatting  in excel (numbers and commas, etc)
for cell in ws["E:E"]:
    cell.number_format = '#,##0'

for cell in ws["D:D"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws["F:F"]:
    cell.number_format = u'"$ "#,##0.00'


wb.save('BoxOfficeReport.xlsx')